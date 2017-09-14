import logging
from datetime import date, datetime
from unittest import skip

from dateutil import relativedelta
import numpy as np
from excel_helper import growth_coefficients
from openpyxl import load_workbook

from excel_helper.helper import ParameterLoader, get_random_variable_definition, DataSeriesLoader, \
    MultiSourceLoader, MCDataset, ExcelSeriesLoaderDataSource, ExcelLoaderDataSource
import pandas as pd

__author__ = 'schien'

import unittest


class TestExcelTool(unittest.TestCase):
    # def setUp(self):
    # self.seq = range(10)

    def test_space_in_var_name(self):
        data = ParameterLoader.from_excel('test.xlsx', size=1, sheet_name='Sheet1')
        res = data['variable with space at end in excel'][0]

        assert res == 2 or res == 4

    def test_open_by_name(self):
        data = ParameterLoader.from_excel('test.xlsx', size=1, sheet_name='Sheet1')
        res = data['c'][0]

        assert (res < 10.) & (res > 3.)

    def test_open_index_zero(self):
        data = ParameterLoader.from_excel('test.xlsx', size=1, sheet_index=0)
        res = data['c'][0]

        assert (res < 10.) & (res > 3.)

    def test_get_row(self):
        data = ParameterLoader.from_excel('test.xlsx', sheet_index=0)

        row = data.get_row('a')
        # print row
        f, p, _ = get_random_variable_definition(row)
        print(p)
        print(f(*p))

    def test_cache(self):
        data = ParameterLoader.from_excel('test.xlsx', size=1, sheet_index=0)
        print(data['a'])

        res = data['a'][0]
        assert res == 4.

    def test_set_scenarios(self):
        data = ParameterLoader.from_excel('test.xlsx', size=1, sheet_index=0)
        res = data['a'][0]

        assert res == 1.

        data.select_scenario('s1')
        res = data['a'][0]

        assert res == 2.

    def test_unset_scenarios(self):
        data = ParameterLoader.from_excel('test.xlsx', size=1, sheet_index=0)
        res = data['a'][0]

        assert res == 1.

        data.select_scenario('s1')
        res = data['a'][0]

        assert res == 2.

        data.unselect_scenario()
        res = data['a'][0]

        assert res == 1.

    def test_cache_with_scenarios(self):
        data = ParameterLoader.from_excel('test.xlsx', size=100000, sheet_index=0)
        a = data['a'][0]
        b = data['b'][0]

        c = a + b
        res = data['a'][0]

        assert res == 1.

        data.select_scenario('s1')
        data['a'][0]
        res = data['a'][0]

        assert res == 2.


class TestDistributionFunctions(unittest.TestCase):
    def test_choice(self):
        data = ParameterLoader.from_excel('test.xlsx', size=1, sheet_index=0)
        res = data['a'][0]
        assert res == 1.

    def test_multiple_choice(self):
        data = ParameterLoader.from_excel('test.xlsx', size=1, sheet_index=0)
        res = data['multiple choice'][0]
        assert res in [1., 2., 3.]

    def test_arrays(self):
        data = ParameterLoader.from_excel('test.xlsx', size=2, sheet_index=0)
        res = data['a']
        # print res
        assert np.array_equal(res, [1., 1.])

    def test_uniform(self):
        data = ParameterLoader.from_excel('test.xlsx', size=1, sheet_index=0)
        res = data['b'][0]

        assert (res < 4.) & (res > 2.)

    def test_triangular(self):
        data = ParameterLoader.from_excel('test.xlsx', size=1, sheet_index=0)
        res = data['c'][0]

        assert (res < 10.) & (res > 3.)


        # def test_module_choice(self):
        # data = ModelLoader('../data/metro_core_network_model_params.xlsx')
        #     res = data['R']
        #     print res


        # def test_module_args(self):
        #     data = ModelLoader('../data/metro_core_network_model_params.xlsx')
        #     eta_M_R = data.get_val('eta_M_R', {'size': 5})  # - $\eta_{M_R}$ metro router overcapacity/ utilisation
        #     print eta_M_R


class TestDataFrameLoader(unittest.TestCase):
    def test_get_dataframe(self):
        # the time axis of our dataset
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')
        # the sample axis our dataset
        samples = 3

        dfl = DataSeriesLoader.from_excel('test.xlsx', times, size=samples, sheet_index=0)
        res = dfl['a']

        # assert np.array_equal(res, [1., 1.])

    def test_metadata(self):
        # the time axis of our dataset
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')
        # the sample axis our dataset
        samples = 3

        dfl = DataSeriesLoader.from_excel('test.xlsx', times, size=samples, sheet_index=0)
        res = dfl['a']

        print(res._metadata)

    def test_from_dataframe(self):
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')

        xls = pd.ExcelFile('test.xlsx')
        df = xls.parse('Sheet1')
        ldr = DataSeriesLoader.from_dataframe(df, times, size=3)
        res = ldr['a']
        print(res)


class TestCAGRCalculation(unittest.TestCase):
    def test_identitical_month(self):
        """
        If start and end are identical, we expect an array of one row of ones of sample size

        :return:
        """
        samples = 3
        alpha = 1  # 100 percent p.a.
        ref_date = date(2009, 1, 1)
        start_date = date(2009, 1, 1)
        end_date = date(2009, 1, 1)

        a = growth_coefficients(start_date, end_date, ref_date, alpha, samples)
        assert np.all(a == np.ones((samples, 1)))

    def test_one_year(self):
        """
        If start and end are one month apart, we expect an array of one row of ones of sample size for the ref month
        and one row with CAGR applied

        :return:
        """
        samples = 3
        alpha = 1  # 100 percent p.a.
        ref_date = date(2009, 1, 1)
        start_date = date(2009, 1, 1)
        end_date = date(2010, 1, 1)

        a = growth_coefficients(start_date, end_date, ref_date, alpha, samples)
        print(a)
        assert np.all(a[0] == np.ones((samples, 1)))
        assert np.all(a[1] == np.ones((samples, 1)) * pow(1 + alpha, 1. / 12))

    def test_one_year(self):
        """
        If start and end are one month apart, we expect an array of one row of ones of sample size for the ref month
        and one row with CAGR applied

        :return:
        """
        samples = 3
        alpha = 0.5  # 100 percent p.a.
        ref_date = date(2009, 1, 1)
        start_date = date(2009, 1, 1)
        end_date = date(2010, 1, 1)

        a = growth_coefficients(start_date, end_date, ref_date, alpha, samples)
        print(a)
        assert np.all(a[0] == np.ones((samples, 1)))
        assert np.all(a[1] == np.ones((samples, 1)) * pow(1 + alpha, 1. / 12))

    def test_negative_growth(self):
        """
        If start and end are one month apart, we expect an array of one row of ones of sample size for the ref month
        and one row with CAGR applied

        :return:
        """
        samples = 3
        alpha = -0.1  # 100 percent p.a.
        ref_date = date(2009, 1, 1)
        start_date = date(2009, 1, 1)
        end_date = date(2010, 1, 1)
        a = growth_coefficients(start_date, end_date, ref_date, alpha, samples)
        print(a)
        assert np.all(a[0] == np.ones((samples, 1)))
        assert np.all(a[-1] == np.ones((samples, 1)) * 1 + alpha)

    def test_refdate_after_start(self):
        """
        If the ref date is greater than the start, we expect an array of one row of ones of sample size for the ref month
        and rows with (1-CAGR)^t applied

        :return:
        """
        samples = 3
        alpha = 0.1  # 10 percent p.a.
        ref_date = date(2009, 2, 1)
        start_date = date(2009, 1, 1)
        end_date = date(2009, 2, 1)

        a = growth_coefficients(start_date, end_date, ref_date, alpha, samples)
        assert a.shape == (2, samples)
        # first row has negative coefficients
        assert np.all(a[0] == np.ones((samples, 1)) * pow(1 - alpha, 1. / 12))
        # second row has ref values
        assert np.all(a[-1] == np.ones((samples, 1)))

    def test_refdate_between_start_and_end(self):
        """
        If the ref date is greater than the start, we expect an array of one row of ones of sample size for the ref month
        and rows with (1-CAGR)^t applied

        :return:
        """
        samples = 3
        alpha = 0.1  # 10 percent p.a.
        ref_date = date(2009, 3, 1)
        start_date = date(2009, 1, 1)
        end_date = date(2009, 6, 1)

        delta = relativedelta.relativedelta(end_date, start_date)
        total_months = delta.months + delta.years * 12 + 1

        ref_delta = relativedelta.relativedelta(ref_date, start_date)
        ref_row_idx = ref_delta.months + ref_delta.years * 12

        a = growth_coefficients(start_date, end_date, ref_date, alpha, samples)
        # print a

        assert a.shape == (total_months, samples)
        # the ref_row_idx has all ones
        assert np.all(a[ref_row_idx] == np.ones((samples, 1)))
        # the row before ref_row_idx has negative coefficients
        assert np.all(a[ref_row_idx - 1] == np.ones((samples, 1)) * pow(1 - alpha, 1. / 12))
        # the row after ref_row_idx has positive coefficients
        assert np.all(a[ref_row_idx + 1] == np.ones((samples, 1)) * pow(1 + alpha, 1. / 12))

        # the last row has positive coefficients
        assert np.all(a[-1] == np.ones((samples, 1)) * pow(1 + alpha, float(total_months - 1 - ref_row_idx) / 12))


class TestDataFrameWithCAGRCalculation(unittest.TestCase):
    def test_simple_CAGR(self):
        """
        Basic test case, applying CAGR to a Pandas Dataframe.

        :return:
        """
        # the time axis of our dataset
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')
        # the sample axis our dataset
        samples = 2

        dfl = DataSeriesLoader.from_excel('test.xlsx', times, size=samples, sheet_index=0)
        res = dfl['a']

        assert res.loc[[datetime(2009, 1, 1)]][0] == 1
        assert np.abs(res.loc[[datetime(2009, 4, 1)]][0] - pow(1.1, 3. / 12)) < 0.00001

    def test_CAGR_ref_date_within_bounds(self):
        """
        Basic test case, applying CAGR to a Pandas Dataframe.

        :return:
        """
        # the time axis of our dataset
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')
        # the sample axis our dataset
        samples = 2

        dfl = DataSeriesLoader.from_excel('test.xlsx', times, size=samples, sheet_index=0)
        res = dfl['a']

        assert res.loc[[datetime(2009, 1, 1)]][0] == 1
        assert np.abs(res.loc[[datetime(2009, 4, 1)]][0] - pow(1.1, 3. / 12)) < 0.00001

    def test_CAGR_ref_date_before_start(self):
        """
        Basic test case, applying CAGR to a Pandas Dataframe.

        :return:
        """
        # the time axis of our dataset
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')
        # the sample axis our dataset
        samples = 2

        dfl = DataSeriesLoader.from_excel('test.xlsx', times, size=samples, sheet_index=0)
        # equivalent to dfl['test_ref_date_before_start']
        self.assertRaises(AssertionError, dfl.__getitem__, 'test_ref_date_before_start')

    def test_CAGR_ref_date_after_end(self):
        """
        Basic test case, applying CAGR to a Pandas Dataframe.

        :return:
        """
        # the time axis of our dataset
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')
        # the sample axis our dataset
        samples = 2

        dfl = DataSeriesLoader.from_excel('test.xlsx', times, size=samples, sheet_index=0)
        # equivalent to dfl['test_ref_date_before_start']
        self.assertRaises(AssertionError, dfl.__getitem__, 'test_ref_date_after_end')

    def test_simple_CAGR_from_pandas(self):
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')

        xls = pd.ExcelFile('test.xlsx')
        df = xls.parse('Sheet1')
        ldr = DataSeriesLoader.from_dataframe(df, times, size=2)
        res = ldr['a']

        assert res.loc[[datetime(2009, 1, 1)]][0] == 1
        assert np.abs(res.loc[[datetime(2009, 4, 1)]][0] - pow(1.1, 3. / 12)) < 0.00001

    def test_simple_CAGR_mm(self):
        """
        Basic test case, applying CAGR to a Pandas Dataframe.

        :return:
        """
        # the time axis of our dataset
        times = pd.date_range('2015-01-01', '2016-01-01', freq='MS')
        # the sample axis our dataset
        samples = 2

        dfl = DataSeriesLoader.from_excel('test.xlsx', times, size=samples, sheet_index=0)
        res = dfl['mm']
        print(res)
        # assert res.loc[[datetime(2009, 1, 1)]][0] == 1
        # assert np.abs(res.loc[[datetime(2009, 4, 1)]][0] - pow(1.1, 3. / 12)) < 0.00001


class TestMultiSourceLoader(unittest.TestCase):
    def test_simple(self):
        data = MultiSourceLoader()

        data.add_source(ExcelLoaderDataSource('test.xlsx', size=1, sheet_index=0))
        # data.add_source(ExcelLoaderDataSource('test.xlsx', size=1, sheet_index=1))
        print(data['a'])
        a = data['a'][0]
        assert a == 1.

        z = data['z'][0]
        assert z == 1.

    @skip
    def test_mean_values(self):
        data = MultiSourceLoader(sample_mean_value=True)

        data.add_source(ExcelLoaderDataSource('test.xlsx', size=10, sheet_index=0))
        data.add_source(ExcelLoaderDataSource('test.xlsx', size=10, sheet_index=1))
        a = data['a'][0]
        assert (a == 1).all()

        z = data['z'][0]
        print(z)
        assert (z == 1.5).all()

    @skip
    def test_mixed_type_multisource(self):
        data = MultiSourceLoader()

        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')
        # the sample axis our dataset
        samples = 2

        data.add_source(ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=0))
        res = data['a']

        assert res.loc[[datetime(2009, 1, 1)]][0] == 1
        assert np.abs(res.loc[[datetime(2009, 4, 1)]][0] - pow(1.1, 3. / 12)) < 0.00001
        data.add_source(ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=1))
        z = data['z'][0]
        assert z >= 1. & z <= 2.

    def test_scenario(self):
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')
        # the sample axis our dataset
        samples = 2

        data = MultiSourceLoader()
        data.add_source(ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=0))

        res = data['a'][0]

        assert res == 1.

        data.set_scenario('s1')
        res = data['a'][0]

        assert res == 2.

        data.reset_scenario()
        res = data['a'][0]

        assert res == 1.

    def test_reload(self):
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')
        # the sample axis our dataset
        samples = 2

        data = MultiSourceLoader()
        data.add_source(ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=0))

        res = data['a'][0]
        assert res == 1.

        wb = load_workbook(filename='test.xlsx')
        ws = wb.worksheets[0]
        ws['E2'] = 4.
        wb.save(filename='test.xlsx')

        data.reload_sources()

        res = data['a'][0]
        assert res == 4.

        wb = load_workbook(filename='test.xlsx')
        ws = wb.worksheets[0]
        ws['E2'] = 1.
        wb.save(filename='test.xlsx')

        data.reload_sources()

        data.set_scenario('s1')
        res = data['a'][0]

        assert res == 2.

        data.reset_scenario()
        res = data['a'][0]

        assert res == 1.


class TestMCDataset(unittest.TestCase):
    def test_simple(self):
        data = MCDataset()
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')
        # the sample axis our dataset
        samples = 2
        data.add_source(ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=0))
        # data.prepare('a')
        res = data['a']
        print(res)
        res = data.b
        print(res)
        # assert res.loc[[datetime(2009, 1, 1)]][0] == 1
        # assert np.abs(res.loc[[datetime(2009, 4, 1)]][0] - pow(1.1, 3. / 12)) < 0.00001

    def test_fix_variables_analytic_mean_choice_distribution(self):
        """
        Test a model with a single random variable - all other variables are fixed to their mean values.
        :return:
        """
        data = MCDataset(single_var='b')
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')

        # the sample axis our dataset
        samples = 10
        data.add_source(ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=0))
        a = data['a']
        print(a)
        assert np.all(np.equal(a, np.ones(a.shape)))

    def test_fix_variables_analytic_mean_uniform_distribution(self):
        """
        Test a model with a single random variable - all other variables are fixed to their mean values.
        :return:
        """
        data = MCDataset(single_var='a')
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')

        # the sample axis our dataset
        samples = 10
        data.add_source(ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=0))
        b = data['b']

        assert np.all(np.equal(b, np.ones(b.shape) * 3))

    def test_fix_variables_analytic_mean_triangular_distribution(self):
        """
        Test a model with a single random variable - all other variables are fixed to their mean values.
        :return:
        """
        data = MCDataset(single_var='a')
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')

        # the sample axis our dataset
        samples = 10
        data.add_source(ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=0))
        c = data['c']

        assert np.all(np.equal(c, np.ones(c.shape) * 19 / 3))

    def test_fix_variables_analytic_mean_normal_distribution(self):
        """
        Test a model with a single random variable - all other variables are fixed to their mean values.
        :return:
        """
        data = MCDataset(single_var='a')
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')

        # the sample axis our dataset
        samples = 10
        data.add_source(ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=0))
        e = data['e']

        assert np.all(np.equal(e, np.ones(e.shape)))

    def test_fix_variables_analytic_mean_normal_zero_distribution(self):
        """
        Test a model with a single random variable - all other variables are fixed to their mean values.
        :return:
        """
        data = MCDataset(single_var='a')
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')

        # the sample axis our dataset
        samples = 10
        data.add_source(ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=0))
        e = data['ez']

        assert np.all(np.equal(e, np.zeros(e.shape)))

    def test_fix_variables_repeat_mean(self):
        """
        Test a model with a single random variable - all other variables are fixed to their mean values.
        :return:
        """
        data = MCDataset(single_var='a')
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')

        # the sample axis our dataset
        samples = 10
        data.add_source(ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=0))

        b = data['b']
        print(b)
        data = MCDataset(single_var='a')
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')

        # the sample axis our dataset
        samples = 10
        data.add_source(ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=0))

        print(data['b'])

        assert np.all(np.equal(b, data['b']))

    def test_fix_variables(self):
        """
        Test a model with a single random variable - all other variables are fixed to their mean values.
        :return:
        """
        data = MCDataset(single_var='b')
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')

        # the sample axis our dataset
        samples = 10
        data.add_source(ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=0))

        a = data['a']
        b = data['b']

        # print(a)
        assert np.all(np.equal(a, np.ones(a.shape)))
        # print(b.mean(dim='samples'))
        c = a * b
        # print(c)
        assert a.shape == c.shape
        d = b * a
        # print(d)
        assert a.shape == d.shape

    def test_single_var_without_cagr(self):
        data = MCDataset(single_var='a', with_single_var_cagr=False)
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')

        # the sample axis our dataset
        samples = 10
        data.add_source(ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=0))

        a = data['a']
        print(a)
        assert np.all(np.equal(a, np.ones(a.shape)))

    def test_single_var_with_cagr(self):
        data = MCDataset(single_var='a')
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')

        # the sample axis our dataset
        samples = 10
        data.add_source(
            ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=0))

        res = data['a']
        print(res)
        january = res.loc[[datetime(2009, 1, 1)]]
        assert np.all(np.equal(january, np.ones(january.shape)))

        april = res.loc[[datetime(2009, 4, 1)]]
        diff = april - np.ones(april.shape) * pow(1.1, 3. / 12)

        assert np.all(np.less(diff, np.ones(april.shape) * 0.00001))

    def test_mean_values_all(self):
        data = MCDataset(sample_mean_value=True)
        times = pd.date_range('2009-01-01', '2009-04-01', freq='MS')
        # the sample axis our dataset
        samples = 3
        data.add_source(ExcelSeriesLoaderDataSource('test.xlsx', times, size=samples, sheet_index=0))

        b = data['b']

        print(b)




        # assert res.loc[[datetime(2009, 1, 1)]][0] == 1
        # assert np.abs(res.loc[[datetime(2009, 4, 1)]][0] - pow(1.1, 3. / 12)) < 0.00001


class TestExcelLoaderMixin(unittest.TestCase):
    def test_repr(self):
        logging.basicConfig(level=logging.INFO)
        # logging.basicConfig(format='%(asctime)s [%(levelname)s] %(message)s', level=logging.INFO)

        source = ExcelLoaderDataSource('test.xlsx', size=1, sheet_index=0)
        print(source)


class KWReceiver(object):
    def __init__(self, **kwargs):
        print(kwargs)


class KWTest(object):
    def __init__(self, **kwargs):
        self.kwargs = kwargs

    def run(self):
        KWReceiver(**self.kwargs)


class TestKWArgs(unittest.TestCase):
    def test_simple(self):
        logging.basicConfig(level=logging.INFO)
        # logging.basicConfig(format='%(asctime)s [%(levelname)s] %(message)s', level=logging.INFO)

        KWTest(a=1, b=2).run()


if __name__ == '__main__':
    unittest.main()