import csv
import importlib
from abc import abstractmethod
from collections import defaultdict
from typing import Dict, List, Set

import numpy as np
import pandas as pd
from dateutil import relativedelta as rdelta
from openpyxl import load_workbook
import logging
from functools import partial

__author__ = 'schien'

param_name_map = {'variable': 'name', 'scenario': 'source_scenarios_string', 'module': 'module_name',
                  'distribution': 'distribution_name', 'param 1': 'param_a', 'param 2': 'param_b', 'param 3': 'param_c',
                  'unit': '', 'CAGR': 'cagr', 'ref date': '', 'label': '', 'tags': '', 'comment': '', 'source': ''}

# logger.basicConfig(level=logger.DEBUG)
logger = logging.getLogger(__name__)


class DistributionFunctionGenerator(object):
    module: str
    distribution: str
    param_a: str
    param_b: str
    param_c: str

    def __init__(self, module_name=None, distribution_name=None, param_a: float = None,
                 param_b: float = None, param_c: float = None, size=None, **kwargs):
        """
        Instantiate a new object.

        :param module_name:
        :param distribution_name:
        :param param_a:
        :param param_b:
        :param param_c:
        :param size:
        :param kwargs: can contain key "sample_mean_value" with bool value
        """
        self.size = size
        self.module_name = module_name
        self.distribution_name = distribution_name
        self.sample_mean_value = kwargs.get('sample_mean_value', False)
        # prepare function arguments
        if distribution_name == 'choice':
            if type(param_a) == str:
                tokens = param_a.split(',')
                params = [float(token.strip()) for token in tokens]
                self.random_function_params = [np.array(params, dtype=np.float)]
            else:
                self.random_function_params = [np.array([i for i in [param_a, param_b, param_c] if i], dtype=np.float)]

            logger.debug(f'setting function params for choice distribution {self.random_function_params}')
        else:
            self.random_function_params = [i for i in [param_a, param_b, param_c] if i not in [None, ""]]

    def get_mean(self, distribution_function):
        """Get the mean value for a distribution.
        If the distribution function is [normal, uniform,choice,triangular] the analytic value is being calculted.
        Else, the distribution is instantiated and then the mean is being calculated.

        :param distribution_function:
        :return: the mean as a scalar
        """
        name = self.distribution_name
        params = self.random_function_params
        if name == 'normal':
            return params[0]
        if name == 'uniform':
            return (params[0] + params[1]) / 2.
        if name == 'choice':
            return params[0].mean()
        if name == 'triangular':
            return (params[0] + params[1] + params[2]) / 3.
        return distribution_function().mean()

    def generate_values(self, *args, **kwargs):
        """
        Generate a sample of values by sampling from a distribution. The size of the sample can be overriden with the 'size' kwarg.

        If `self.sample_mean_value == True` the sample will contain "size" times the mean value.

        :param args:
        :param kwargs:
        :return: sample as vector of given size
        """
        sample_size = kwargs.get('size', self.size)

        f = self.instantiate_distribution_function(self.module_name, self.distribution_name)
        distribution_function = partial(f, *self.random_function_params, size=sample_size)

        if self.sample_mean_value:
            sample = np.full(sample_size, self.get_mean(distribution_function))
        else:
            sample = distribution_function()

        return sample

    @staticmethod
    def instantiate_distribution_function(module_name, distribution_name):
        module = importlib.import_module(module_name)
        func = getattr(module, distribution_name)
        return func


class Parameter(object):
    """
    A single parameter
    """

    name: str
    unit: str
    comment: str
    source: str
    scenario: str

    processes: Dict[str, List]

    "optional comma-separated list of tags"
    tags: str

    def __init__(self, name, tags=None, source_scenarios_string: str = None, unit: str = None,
                 comment: str = None, source: str = None,
                 **kwargs):
        # The source definition of scenarios. A comma-separated list
        self.source = source
        self.comment = comment

        self.unit = unit
        self.source_scenarios_string = source_scenarios_string
        self.tags = tags
        self.name = name

        self.scenario = None
        self.cache = None

        # track the usages of this parameter per process as a list of process-specific variable names that are backed by this parameter
        self.processes = defaultdict(list)

        self.kwargs = kwargs

    def __call__(self, settings=None, *args, **kwargs):
        """
        Returns the same value every time called.
        :param args:
        :param kwargs:
        :return:
        """
        if self.cache is None:
            kwargs['name'] = self.name
            kwargs['unit'] = self.unit
            kwargs['tags'] = self.tags
            kwargs['scenario'] = self.scenario

            if not settings:
                settings = {}

            common_args = {'size': settings.get('sample_size', 1),
                           'sample_mean_value': settings.get('sample_mean_value', False)}
            common_args.update(**self.kwargs)

            if settings.get('use_time_series', False):
                generator = ExponentialGrowthTimeSeriesGenerator(**common_args, times=settings['times'])
            else:
                generator = DistributionFunctionGenerator(**common_args)

            self.cache = generator.generate_values(*args, **kwargs)
        return self.cache

    def add_usage(self, process_name, variable_name):
        # add the name of a variable of a process model that is backed by this parameter
        self.processes[process_name].append(variable_name)


class ExponentialGrowthTimeSeriesGenerator(DistributionFunctionGenerator):
    cagr: str
    ref_date: str

    def __init__(self, cagr=None, times=None, size=None, index_names=None, ref_date=None, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.cagr = cagr if cagr else 0
        self.ref_date = ref_date
        self.times = times
        self.size = size
        iterables = [times, range(0, size)]
        self._multi_index = pd.MultiIndex.from_product(iterables, names=index_names)
        assert type(times.freq) == pd.tseries.offsets.MonthBegin, 'Time index must have monthly frequency'

    def generate_values(self, *args, **kwargs):
        """
        Instantiate a random variable and apply annual growth factors.

        :return:
        """
        values = super().generate_values(*args, **kwargs, size=(len(self.times) * self.size,))
        alpha = self.cagr

        # @todo - fill to cover the entire time: define rules for filling first
        ref_date = self.ref_date if self.ref_date else self.times[0].to_pydatetime()
        assert ref_date >= self.times[0].to_pydatetime(), 'Ref date must be within variable time span.'
        assert ref_date <= self.times[-1].to_pydatetime(), 'Ref date must be within variable time span.'

        start_date = self.times[0].to_pydatetime()
        end_date = self.times[-1].to_pydatetime()

        a = growth_coefficients(start_date, end_date, ref_date, alpha, self.size)

        values *= a.ravel()

        df = pd.DataFrame(values)
        df.columns = [kwargs['name']]
        df.set_index(self._multi_index, inplace=True)
        # @todo this is a hack to return a series with index as I don't know how to set an index and rename a series
        data_series = df.iloc[:, 0]
        data_series._metadata = kwargs
        data_series.index.rename(['time', 'samples'], inplace=True)

        return data_series


def growth_coefficients(start_date, end_date, ref_date, alpha, samples):
    """
    Build a matrix of growth factors according to the CAGR formula  y'=y0 (1+a)^(t'-t0).
    The
    """
    if ref_date < start_date:
        raise ValueError("Ref data must be >= start date.")
    if ref_date > end_date:
        raise ValueError("Ref data must be >= start date.")
    if ref_date > start_date and alpha >= 1:
        raise ValueError("For a CAGR >= 1, ref date and start date must be the same.")
    # relative delta will be positive if ref date is >= start date (which it should with above assertions)
    delta_ar = rdelta.relativedelta(ref_date, start_date)
    ar = delta_ar.months + 12 * delta_ar.years
    delta_br = rdelta.relativedelta(end_date, ref_date)
    br = delta_br.months + 12 * delta_br.years

    # we place the ref point on the lower interval (delta_ar + 1) but let it start from 0
    # in turn we let the upper interval start from 1
    g = np.fromfunction(lambda i, j: np.power(1 - alpha, np.abs(i) / 12), (ar + 1, samples), dtype=float)
    h = np.fromfunction(lambda i, j: np.power(1 + alpha, np.abs(i + 1) / 12), (br, samples), dtype=float)
    g = np.flipud(g)
    # now join the two arrays
    return np.vstack((g, h))


class ParameterScenarioSet(object):
    """
    The set of all version of a parameter for all the scenarios.
    """
    default_scenario = 'default'

    "the name of the parameters in this set"
    parameter_name: str
    scenarios: Dict[str, Parameter]

    def __init__(self):
        self.scenarios = {}

    def add_scenario(self, parameter: 'Parameter', scenario_name: str = default_scenario):
        """
        Add a scenario for this parameter.

        :param scenario_name:
        :param parameter:
        :return:
        """
        self.scenarios[scenario_name] = parameter

    def __getitem__(self, item):
        return self.scenarios.__getitem__(item)

    def __setitem__(self, key, value):
        return self.scenarios.__setitem__(key, value)


class ParameterRepository(object):
    """
    Contains all known parameter definitions (so that it is not necessary to re-read the excel file for repeat param accesses).
    The param definitions are independent from the sampling (the Param.__call__ method). Repeat access to __call__ will
    create new samples.

    Internally, parameters are organised together with all the scenario variants in a single ParameterScenarioSet.

    """
    parameter_sets: Dict[str, ParameterScenarioSet]
    tags: Dict[str, Dict[str, Set[Parameter]]]

    def __init__(self):
        self.parameter_sets = defaultdict(ParameterScenarioSet)
        self.tags = defaultdict(lambda: defaultdict(set))

    def add_all(self, parameters: List[Parameter]):
        for p in parameters:
            self.add_parameter(p)

    def add_parameter(self, parameter: Parameter):
        """
        A parameter can have several scenarios. They are specified as a comma-separated list in a string.
        :param parameter:
        :return:
        """

        # try reading the scenarios from the function arg or from the parameter attribute
        scenario_string = parameter.source_scenarios_string
        if scenario_string:
            _scenarios = [i.strip() for i in scenario_string.split(',')]
            self.fill_missing_attributes_from_default_parameter(parameter)

        else:
            _scenarios = [ParameterScenarioSet.default_scenario]

        for scenario in _scenarios:
            parameter.scenario = scenario
            self.parameter_sets[parameter.name][scenario] = parameter

        # record all tags for this parameter
        if parameter.tags:
            _tags = [i.strip() for i in parameter.tags.split(',')]
            for tag in _tags:
                self.tags[tag][parameter.name].add(parameter)

    def fill_missing_attributes_from_default_parameter(self, param):
        """
        Empty fields in Parameter definitions in scenarios are populated with default values.

        E.g. in the example below, the source for the Power_TV variable in the 8K scenario would also be EnergyStar.

        | name     | scenario | val | tags   | source     |
        |----------|----------|-----|--------|------------|
        | Power_TV |          | 60  | UD, TV | EnergyStar |
        | Power_TV | 8K       | 85  | new_tag|            |

        **Note** tags must not differ. In the example above, the 8K scenario variable the tags value would be overwritten
        with the default value.

        :param param:
        :return:
        """
        if not self.exists(param.name) or not ParameterScenarioSet.default_scenario in self.parameter_sets[
            param.name].scenarios.keys():
            logger.warning(
                f'No default value for param {param.name} found.')
            return
        default = self.parameter_sets[param.name][ParameterScenarioSet.default_scenario]
        for att_name, att_value in default.__dict__.items():
            if att_name in ['unit', 'label', 'comment', 'source', 'tags']:

                if att_name == 'tags' and default.tags != param.tags:
                    logger.warning(
                        f'For param {param.name} for scenarios {param.source_scenarios_string}, tags is different from default parameter tags. Overwriting with default values.')
                    setattr(param, att_name, att_value)

                if not getattr(param, att_name):
                    logger.info(
                        f'For param {param.name} for scenarios {param.source_scenarios_string}, populating attribute {att_name} with value {att_value} from default parameter.')

                    setattr(param, att_name, att_value)

    def __getitem__(self, item) -> Parameter:
        """
        Return the default scenario parameter for a given variable name
        :param item: the name of the variable
        :return:
        """
        return self.get_parameter(item, scenario_name=ParameterScenarioSet.default_scenario)

    def get_parameter(self, param_name, scenario_name=None) -> Parameter:
        scenario = scenario_name if scenario_name else ParameterScenarioSet.default_scenario
        try:
            return self.parameter_sets[param_name][scenario]
        except KeyError:
            raise KeyError(f"{param_name} for scenario {scenario} not found")

    def find_by_tag(self, tag) -> Dict[str, Set[Parameter]]:
        """
        Get all registered dicts that are registered for a tag

        :param tag: str - single tag
        :return: a dict of {param name: set[Parameter]} that contains all ParameterScenarioSets for
        all parameter names with a given tag
        """
        return self.tags[tag]

    def exists(self, param):
        return param in self.parameter_sets.keys()


class ExcelHandler(object):
    @abstractmethod
    def load_definitions(self, sheet_name, filename=None):
        raise NotImplementedError()


class OpenpyxlExcelHandler(ExcelHandler):
    def load_definitions(self, sheet_name, filename=None):
        definitions = []
        wb = load_workbook(filename=filename, data_only=True)
        _sheet_names = [sheet_name] if sheet_name else wb.sheetnames
        for _sheet_name in _sheet_names:
            sheet = wb.get_sheet_by_name(_sheet_name)
            rows = list(sheet.rows)
            header = [cell.value for cell in rows[0]]

            if header[0] != 'variable':
                continue

            for row in rows[1:]:
                values = {}
                for key, cell in zip(header, row):
                    values[key] = cell.value
                definitions.append(values)
        return definitions


class Xlsx2CsvHandler(ExcelHandler):
    def load_definitions(self, sheet_name, filename=None):
        from xlsx2csv import Xlsx2csv
        data = Xlsx2csv(filename, inmemory=True).convert(None, sheetid=0)

        definitions = []

        _sheet_names = [sheet_name] if sheet_name else [data.keys()]

        for _sheet_name in _sheet_names:
            sheet = data[_sheet_name]

            header = sheet.header
            if header[0] != 'variable':
                continue

            for row in sheet.rows:
                values = {}
                for key, cell in zip(header, row):
                    values[key] = cell
                definitions.append(values)
        return definitions


class CSVHandler(ExcelHandler):
    def load_definitions(self, sheet_name, filename=None):
        return csv.DictReader(open(filename), delimiter=',')


class XLRDExcelHandler(ExcelHandler):
    @staticmethod
    def get_sheet_range_bounds(filename, sheet_name):
        import xlrd
        wb = xlrd.open_workbook(filename)
        sheet = wb.sheet_by_name(sheet_name)
        rows = list(sheet.get_rows())
        return len(rows)

    def load_definitions(self, sheet_name, filename=None):
        import xlrd
        wb = xlrd.open_workbook(filename)
        sh = None

        definitions = []

        _sheet_names = [sheet_name] if sheet_name else [sh.name for sh in wb.sheets()]

        for _sheet_name in _sheet_names:
            sheet = wb.sheet_by_name(_sheet_name)
            rows = list(sheet.get_rows())
            header = [cell.value for cell in rows[0]]

            if header[0] != 'variable':
                continue

            for row in rows[1:]:
                values = {}
                for key, cell in zip(header, row):
                    values[key] = cell.value
                definitions.append(values)
        return definitions


class XLWingsExcelHandler(ExcelHandler):
    def load_definitions(self, sheet_name, filename=None):
        import xlwings as xw
        definitions = []
        wb = xw.Book(fullname=filename)
        _sheet_names = [sheet_name] if sheet_name else wb.sheets
        for _sheet_name in _sheet_names:
            sheet = wb.sheets[_sheet_name]
            range = sheet.range('A1').expand()
            rows = range.rows
            header = [cell.value for cell in rows[0]]

            # check if this sheet contains parameters or if it documentation
            if header[0] != 'variable':
                continue

            total_rows = XLRDExcelHandler.get_sheet_range_bounds(filename, _sheet_name)
            range = sheet.range((1, 1), (total_rows, len(header)))
            rows = range.rows
            for row in rows[1:]:
                values = {}
                for key, cell in zip(header, row):
                    values[key] = cell.value
                definitions.append(values)
        return definitions


class ExcelParameterLoader(object):
    """Utility to populate ParameterRepository from spreadsheets.

        The structure of the spreadsheets is:

        | variable | ... |
        |----------|-----|
        |   ...    | ... |

        If the first row in a spreadsheet does not contain they keyword 'variable' the sheet is ignored.

       """

    def __init__(self, filename, excel_handler='xlrd', **kwargs):
        self.filename = filename

        logger.info(f'Using {excel_handler} excel handler')
        excel_handler_instance = None
        if excel_handler == 'openpyxl':
            excel_handler_instance = OpenpyxlExcelHandler()
        if excel_handler == 'xlsx2csv':
            excel_handler_instance = Xlsx2CsvHandler()
        if excel_handler == 'xlwings':
            excel_handler_instance = XLWingsExcelHandler()
        if excel_handler == 'xlrd':
            excel_handler_instance = XLRDExcelHandler()

        self.excel_handler: ExcelHandler = excel_handler_instance

    def load_parameter_definitions(self, sheet_name: str = None):
        """
        Load variable text from rows in excel file.
        If no spreadsheet arg is given, all spreadsheets are loaded.
        The first cell in the first row in a spreadsheet must contain the keyword 'variable' or the sheet is ignored.

        Any cells used as titles (with no associated value) are also added to the returned dictionary. However, the
        values associated with each header will be None. For example, given the speadsheet:

        | variable | A | B |
        |----------|---|---|
        | Title    |   |   |
        | Entry    | 1 | 2 |

        The following list of definitions would be returned:

        [ { variable: 'Title', A: None, B: None }
        , { variable: 'Entry', A: 1   , B: 2    }
        ]

        :param sheet_name:
        :return: list of dicts with {header col name : cell value} pairs
        """
        definitions = self.excel_handler.load_definitions(sheet_name, filename=self.filename)

        return definitions

    def load_into_repo(self, repository: ParameterRepository = None, sheet_name: str = None):
        """
        Create a Repo from an excel file.
        :param repository: the repository to load into
        :param sheet_name:
        :return:
        """
        repository.add_all(self.load_parameters(sheet_name))

    def load_parameters(self, sheet_name):

        parameter_definitions = self.load_parameter_definitions(sheet_name=sheet_name)
        params = []
        for _def in parameter_definitions:

            # substitute names from the headers with the kwargs names in the Parameter and Distributions classes
            # e.g. 'variable' -> 'name', 'module' -> 'module_name', etc
            parameter_kwargs_def = {}
            for k, v in _def.items():
                if k in param_name_map:
                    if param_name_map[k]:
                        parameter_kwargs_def[param_name_map[k]] = v
                    else:
                        parameter_kwargs_def[k] = v

            name_ = parameter_kwargs_def['name']
            del parameter_kwargs_def['name']
            p = Parameter(name_, **parameter_kwargs_def)
            params.append(p)
        return params
